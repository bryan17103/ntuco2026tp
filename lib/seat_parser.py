from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# ===== 座位區範圍 =====
SEAT_START_COL = column_index_from_string("E")
SEAT_END_COL = column_index_from_string("AT")

# 左右排號欄
LEFT_ROW_LABEL_COL = column_index_from_string("B")
RIGHT_ROW_LABEL_COL = column_index_from_string("AW")

# 右側 legend 欄位
LEGEND_COLOR_COL = column_index_from_string("AX")
LEGEND_LABEL_COL = column_index_from_string("AY")


def get_fill_color(cell):
    fill = cell.fill
    if not fill:
        return None

    color = fill.fgColor
    if not color:
        return None

    if color.type == "rgb":
        return color.rgb
    elif color.type == "indexed":
        return f"indexed:{color.indexed}"
    elif color.type == "theme":
        return f"theme:{color.theme}"

    return None


def normalize_row_label(value):

    if value is None:
        return ""

    if isinstance(value, (int, float)):
        if int(value) == value:
            return str(int(value))
        return str(value)

    return str(value).strip()


def label_to_zone_price_available(label):
  
    text = str(label).strip()

    # 特殊席
    if "工作席" in text:
        return "staff", 0, False

    if "攝影" in text:
        return "camera", 0, False

    if "貴賓" in text:
        return "vip", 0, False

    if "輪椅陪同" in text:
        return "companion", 300, False

    if "輪椅" in text:
        return "wheelchair", 300, False

    # 團內票區
    if "團內" in text:
        if "500" in text:
            return "group-500", 400, True
        if "300" in text:
            return "group-300", 240, True
        if "200" in text:
            return "group-200", 160, True

    # 一般票區
    if "500" in text:
        return "regular-500", 500, False
    if "300" in text:
        return "regular-300", 300, False
    if "200" in text:
        return "regular-200", 200, False

    return "unknown", 0, False


def build_legend_label_map(ws):

    legend_map = {}

    for row in range(1, ws.max_row + 1):
        color_cell = ws.cell(row, LEGEND_COLOR_COL)
        label_cell = ws.cell(row, LEGEND_LABEL_COL)

        color = get_fill_color(color_cell)
        label = label_cell.value

        if not color or label is None:
            continue

        label = str(label).strip()
        if not label:
            continue

        legend_map[color] = label

    return legend_map


def build_color_map(ws):

    legend_label_map = build_legend_label_map(ws)
    color_map = {}

    for color, label in legend_label_map.items():
        zone, price, available = label_to_zone_price_available(label)
        color_map[color] = (zone, price, available)

    return color_map


def parse_seat_map(filepath, debug=False):

    wb = load_workbook(filepath)
    ws = wb.active

    color_map = build_color_map(ws)

    seats = []
    row_labels = {}

    for excel_row in range(1, ws.max_row + 1):
        left_label = normalize_row_label(ws.cell(excel_row, LEFT_ROW_LABEL_COL).value)
        right_label = normalize_row_label(ws.cell(excel_row, RIGHT_ROW_LABEL_COL).value)

        # 左側優先，沒有再用右側
        row_label = left_label or right_label

        if row_label:
            row_labels[excel_row] = row_label

        for excel_col in range(SEAT_START_COL, SEAT_END_COL + 1):
            cell = ws.cell(excel_row, excel_col)
            value = cell.value

            # 只有數字格才視為座位
            if isinstance(value, (int, float)):
                color = get_fill_color(cell)
                zone, price, available = color_map.get(color, ("unknown", 0, False))

                if debug and color not in color_map:
                    print(
                        f"⚠️ 未定義顏色: row={excel_row}, col={excel_col}, "
                        f"seat={value}, color={color}"
                    )

                seats.append({
                    "seat_number": int(value),
                    "excel_row": excel_row,
                    "excel_col": excel_col,
                    "row_label": row_labels.get(excel_row, ""),
                    "zone": zone,
                    "price": price,
                    "color": color,
                    "available": available,
                })

    return seats, row_labels, color_map